"""
Generate 30 classic .xlsx files for testing Excel-to-PDF conversion.
Each file corresponds to a test case in ClassicExcelToPdfTests.cs.

Usage:
    pip install openpyxl
    python generate_classic_xlsx.py

Output directory: ./output/
"""

import os
import string
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")


def ensure_output_dir():
    os.makedirs(OUTPUT_DIR, exist_ok=True)


def save(wb: Workbook, filename: str):
    path = os.path.join(OUTPUT_DIR, filename)
    wb.save(path)
    print(f"  ✔ {filename}")


# ── 01. Basic table with headers ────────────────────────────────────────
def classic01_basic_table_with_headers():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Age", "City"])
    ws.append(["Alice", 30, "New York"])
    ws.append(["Bob", 25, "London"])
    ws.append(["Charlie", 35, "Tokyo"])
    ws.append(["Diana", 28, "Paris"])
    save(wb, "classic01_basic_table_with_headers.xlsx")


# ── 02. Multiple worksheets ─────────────────────────────────────────────
def classic02_multiple_worksheets():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sales"
    ws1.append(["Quarter", "Revenue"])
    ws1.append(["Q1", 100])
    ws1.append(["Q2", 200])
    ws1.append(["Q3", 350])
    ws1.append(["Q4", 480])

    ws2 = wb.create_sheet("Costs")
    ws2.append(["Category", "Amount"])
    ws2.append(["Rent", 500])
    ws2.append(["Salary", 3000])
    ws2.append(["Utilities", 200])

    ws3 = wb.create_sheet("Summary")
    ws3.append(["Metric", "Value"])
    ws3.append(["Total Revenue", 1130])
    ws3.append(["Total Costs", 3700])
    ws3.append(["Net", -2570])
    save(wb, "classic02_multiple_worksheets.xlsx")


# ── 03. Empty workbook (no data rows) ───────────────────────────────────
def classic03_empty_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # No data at all
    save(wb, "classic03_empty_workbook.xlsx")


# ── 04. Single cell ─────────────────────────────────────────────────────
def classic04_single_cell():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Hello"
    save(wb, "classic04_single_cell.xlsx")


# ── 05. Wide table (26 columns A–Z) ─────────────────────────────────────
def classic05_wide_table():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = list(string.ascii_uppercase)  # A-Z
    ws.append(headers)
    for row_idx in range(1, 6):
        ws.append([f"{ch}{row_idx}" for ch in headers])
    save(wb, "classic05_wide_table.xlsx")


# ── 06. Tall table (200 rows → multi-page) ──────────────────────────────
def classic06_tall_table():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Row#", "Value", "Description"])
    for i in range(1, 201):
        ws.append([f"Row{i}", f"Val{i}", f"This is the description for row number {i}"])
    save(wb, "classic06_tall_table.xlsx")


# ── 07. Numbers only ────────────────────────────────────────────────────
def classic07_numbers_only():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append([1.0, 2.0, 3.0])
    ws.append([4.0, 5.0, 6.0])
    ws.append([7.0, 8.0, 9.0])
    ws.append([10.0, 100.0, 1000.0])
    save(wb, "classic07_numbers_only.xlsx")


# ── 08. Mixed text and numbers ──────────────────────────────────────────
def classic08_mixed_text_and_numbers():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Item", "Amount"])
    ws.append(["Item", 10.5])
    ws.append(["Tax", 0.08])
    ws.append(["Total", 10.58])
    ws.append(["Discount", -1.5])
    ws.append(["Final", 9.08])
    save(wb, "classic08_mixed_text_and_numbers.xlsx")


# ── 09. Long text content ───────────────────────────────────────────────
def classic09_long_text():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Long Text Column"])
    ws.append(["X" * 500])
    ws.append(["A" * 300 + " " + "B" * 200])
    ws.append(["Short"])
    ws.append(["Y" * 1000])
    save(wb, "classic09_long_text.xlsx")


# ── 10. Special XML characters ──────────────────────────────────────────
def classic10_special_xml_characters():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Special Characters"])
    ws.append(["A&B"])
    ws.append(["<tag>"])
    ws.append(['"quoted"'])
    ws.append(["it's"])
    ws.append(["Tom & Jerry < Batman > Superman"])
    ws.append(['He said "hello" & she replied \'hi\''])
    save(wb, "classic10_special_xml_characters.xlsx")


# ── 11. Sparse rows (gaps between data rows) ────────────────────────────
def classic11_sparse_rows():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value="First")
    ws.cell(row=5, column=1, value="Fifth")
    ws.cell(row=10, column=1, value="Tenth")
    ws.cell(row=20, column=1, value="Twentieth")
    ws.cell(row=50, column=1, value="Fiftieth")
    save(wb, "classic11_sparse_rows.xlsx")


# ── 12. Sparse columns (A, D filled; B, C empty) ───────────────────────
def classic12_sparse_columns():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Left"
    ws["D1"] = "Right"
    ws["A2"] = "Data1"
    ws["F2"] = "FarRight"
    ws["A3"] = "Row3"
    ws["J3"] = "VeryFar"
    save(wb, "classic12_sparse_columns.xlsx")


# ── 13. Date-like strings ───────────────────────────────────────────────
def classic13_date_strings():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Date", "Event"])
    ws.append(["2025-01-15", "Launch"])
    ws.append(["2025-06-30", "Release"])
    ws.append(["2025-12-25", "Holiday"])
    ws.append(["2026-01-01", "New Year"])
    ws.append(["2026-02-23", "Today"])
    save(wb, "classic13_date_strings.xlsx")


# ── 14. Decimal numbers ─────────────────────────────────────────────────
def classic14_decimal_numbers():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Constant", "Value"])
    ws.append(["Pi", 3.14159])
    ws.append(["e", 2.71828])
    ws.append(["Sqrt(2)", 1.41421])
    ws.append(["Phi", 1.61803])
    ws.append(["Ln(2)", 0.69315])
    save(wb, "classic14_decimal_numbers.xlsx")


# ── 15. Negative numbers ────────────────────────────────────────────────
def classic15_negative_numbers():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Label", "Value"])
    ws.append(["Loss", -100.0])
    ws.append(["Small Loss", -0.5])
    ws.append(["Zero", 0.0])
    ws.append(["Gain", 50.0])
    ws.append(["Big Loss", -99999.99])
    ws.append(["Tiny", -0.001])
    save(wb, "classic15_negative_numbers.xlsx")


# ── 16. Percentage-like strings ──────────────────────────────────────────
def classic16_percentage_strings():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Metric", "Rate"])
    ws.append(["Conversion", "12.5%"])
    ws.append(["Bounce", "45.3%"])
    ws.append(["Retention", "88.7%"])
    ws.append(["Churn", "3.2%"])
    ws.append(["Growth", "156.0%"])
    save(wb, "classic16_percentage_strings.xlsx")


# ── 17. Currency-like strings ───────────────────────────────────────────
def classic17_currency_strings():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Item", "Price"])
    ws.append(["Widget", "$19.99"])
    ws.append(["Gadget", "$149.00"])
    ws.append(["Premium", "$1,299.99"])
    ws.append(["Budget", "$4.50"])
    ws.append(["Euro Item", "€49.99"])
    ws.append(["Yen Item", "¥5000"])
    save(wb, "classic17_currency_strings.xlsx")


# ── 18. Large dataset (1000 rows × 10 cols) ─────────────────────────────
def classic18_large_dataset():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = [f"Col{c}" for c in range(10)]
    ws.append(headers)
    for r in range(1000):
        ws.append([f"R{r}C{c}" for c in range(10)])
    save(wb, "classic18_large_dataset.xlsx")


# ── 19. Single column list ──────────────────────────────────────────────
def classic19_single_column_list():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Items"])
    for i in range(1, 21):
        ws.append([f"Item {i}"])
    save(wb, "classic19_single_column_list.xlsx")


# ── 20. All empty cells ─────────────────────────────────────────────────
def classic20_all_empty_cells():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["", "", ""])
    ws.append(["", "", ""])
    ws.append(["", "", ""])
    save(wb, "classic20_all_empty_cells.xlsx")


# ── 21. Header only (no data rows) ──────────────────────────────────────
def classic21_header_only():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Col1", "Col2", "Col3", "Col4", "Col5"])
    save(wb, "classic21_header_only.xlsx")


# ── 22. Very long sheet name ────────────────────────────────────────────
def classic22_long_sheet_name():
    wb = Workbook()
    # Excel sheet name max is 31 characters
    ws = wb.active
    ws.title = "VeryLongSheetNameThatIsMaxLen"
    ws.append(["Data", "Value"])
    ws.append(["Row1", 100])
    ws.append(["Row2", 200])
    save(wb, "classic22_long_sheet_name.xlsx")


# ── 23. Unicode / CJK text ──────────────────────────────────────────────
def classic23_unicode_text():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Language", "Greeting", "Extra"])
    ws.append(["English", "Hello", "World"])
    ws.append(["Chinese", "你好", "世界"])
    ws.append(["Japanese", "こんにちは", "世界"])
    ws.append(["Korean", "안녕하세요", "세계"])
    ws.append(["Arabic", "مرحبا", "العالم"])
    ws.append(["Emoji", "😀🎉", "✅❌"])
    save(wb, "classic23_unicode_text.xlsx")


# ── 24. Red text (colored) ──────────────────────────────────────────────
def classic24_red_text():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    red_font = Font(color="FF0000", size=11)
    normal_font = Font(size=11)

    ws.append(["Status", "Message"])
    cell_a = ws.cell(row=2, column=1, value="Error")
    cell_a.font = red_font
    cell_b = ws.cell(row=2, column=2, value="Something went wrong")
    cell_b.font = red_font

    cell_a2 = ws.cell(row=3, column=1, value="OK")
    cell_a2.font = normal_font
    cell_b2 = ws.cell(row=3, column=2, value="All systems operational")
    cell_b2.font = normal_font

    cell_a3 = ws.cell(row=4, column=1, value="Warning")
    cell_a3.font = Font(color="FFA500", size=11)
    cell_b3 = ws.cell(row=4, column=2, value="Check disk space")
    cell_b3.font = Font(color="FFA500", size=11)
    save(wb, "classic24_red_text.xlsx")


# ── 25. Multiple colors ─────────────────────────────────────────────────
def classic25_multiple_colors():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    colors = [
        ("Red", "FF0000"),
        ("Green", "00FF00"),
        ("Blue", "0000FF"),
        ("Yellow", "FFFF00"),
        ("Magenta", "FF00FF"),
        ("Cyan", "00FFFF"),
        ("Orange", "FFA500"),
        ("Purple", "800080"),
    ]
    ws.append(["Color Name", "Sample Text"])
    for name, color_hex in colors:
        row = ws.max_row + 1
        cell_a = ws.cell(row=row, column=1, value=name)
        cell_a.font = Font(color=color_hex, size=11)
        cell_b = ws.cell(row=row, column=2, value=f"This is {name.lower()} text")
        cell_b.font = Font(color=color_hex, size=11)
    save(wb, "classic25_multiple_colors.xlsx")


# ── 26. Inline strings ──────────────────────────────────────────────────
def classic26_inline_strings():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Inline1", "Inline2", "Inline3"])
    ws.append(["ValueA", "ValueB", "ValueC"])
    ws.append(["Test1", "Test2", "Test3"])
    save(wb, "classic26_inline_strings.xlsx")


# ── 27. Single row (horizontal data) ────────────────────────────────────
def classic27_single_row():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"])
    save(wb, "classic27_single_row.xlsx")


# ── 28. Duplicate values ────────────────────────────────────────────────
def classic28_duplicate_values():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Yes", "No", "Yes", "No"])
    ws.append(["No", "Yes", "No", "Yes"])
    ws.append(["Yes", "Yes", "Yes", "Yes"])
    ws.append(["No", "No", "No", "No"])
    ws.append(["Yes", "No", "Yes", "No"])
    save(wb, "classic28_duplicate_values.xlsx")


# ── 29. Formula-result values ────────────────────────────────────────────
def classic29_formula_results():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["A", "B", "Sum", "Product"])
    ws.append([10, 20, "=A2+B2", "=A2*B2"])
    ws.append([5, 15, "=A3+B3", "=A3*B3"])
    ws.append([100, 200, "=A4+B4", "=A4*B4"])
    ws.append(["", "", "=SUM(C2:C4)", "=SUM(D2:D4)"])
    save(wb, "classic29_formula_results.xlsx")


# ── 30. Mixed empty and filled sheets ───────────────────────────────────
def classic30_mixed_empty_and_filled_sheets():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Empty"
    # No data in first sheet

    ws2 = wb.create_sheet("Data")
    ws2.append(["Hello", "World"])
    ws2.append(["Foo", "Bar"])
    ws2.append(["Baz", "Qux"])

    ws3 = wb.create_sheet("AlsoEmpty")
    # No data in third sheet

    ws4 = wb.create_sheet("MoreData")
    ws4.append(["Column1", "Column2", "Column3"])
    ws4.append([1, 2, 3])
    save(wb, "classic30_mixed_empty_and_filled_sheets.xlsx")


# ── Main ─────────────────────────────────────────────────────────────────
def main():
    ensure_output_dir()
    print(f"Generating 30 classic .xlsx files in: {OUTPUT_DIR}\n")

    generators = [
        classic01_basic_table_with_headers,
        classic02_multiple_worksheets,
        classic03_empty_workbook,
        classic04_single_cell,
        classic05_wide_table,
        classic06_tall_table,
        classic07_numbers_only,
        classic08_mixed_text_and_numbers,
        classic09_long_text,
        classic10_special_xml_characters,
        classic11_sparse_rows,
        classic12_sparse_columns,
        classic13_date_strings,
        classic14_decimal_numbers,
        classic15_negative_numbers,
        classic16_percentage_strings,
        classic17_currency_strings,
        classic18_large_dataset,
        classic19_single_column_list,
        classic20_all_empty_cells,
        classic21_header_only,
        classic22_long_sheet_name,
        classic23_unicode_text,
        classic24_red_text,
        classic25_multiple_colors,
        classic26_inline_strings,
        classic27_single_row,
        classic28_duplicate_values,
        classic29_formula_results,
        classic30_mixed_empty_and_filled_sheets,
    ]

    for gen in generators:
        gen()

    print(f"\nDone! {len(generators)} files generated.")


if __name__ == "__main__":
    main()
